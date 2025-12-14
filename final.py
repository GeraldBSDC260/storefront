import openpyxl
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.chart import BarChart, LineChart, Reference

# Function: createChart
# Description: Reads an Excel file, extracts data, saves to final.xlsx, and creates a chart.
# Arguments:
#   file_path (str): The full path to your source Excel file.
#   chart_type (str): The type of chart ('line' or 'bar').
# Returns: None
def createChart(file_path, chart_type):
    # Requirement: Ask user to choose the data source
    print("\n--- Select Data Source ---")
    print("1. Initial Data (e.g., Pounds, Inches, Fahrenheit)")
    print("2. Converted Data (e.g., Kilograms, Centimeters, Celsius)")
    choice = input("Enter selection (1 or 2): ")

    # Excel indices: A=0, B=1, C=2. We assume Date=A, Initial=B, Converted=C.
    data_col_index = 2 if choice == '2' else 1
    y_axis_label = "Converted Units" if choice == '2' else "Initial Units"

    extracted_data = []
    try:
        # data_only=True ensures we get the calculated value, not the formula
        source_wb = load_workbook(file_path, data_only=True)
        source_ws = source_wb.active
        
        # Requirement: Open file, extract data, and cast types
        # iter_rows skips the header (row 1)
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                try:
                    date_val = row[0]
                    numeric_val = float(row[data_col_index])
                    extracted_data.append([date_val, numeric_val])
                except (ValueError, TypeError, IndexError):
                    continue
    except FileNotFoundError:
        print(f"\nERROR: Could not find file at: {file_path}")
        return
    except Exception as e:
        print(f"\nAN ERROR OCCURRED: {e}")
        return

    # Requirement: Save data to final.xlsx
    final_wb = Workbook()
    final_ws = final_wb.active
    final_ws.title = "Summary"
    final_ws.append(["Date", y_axis_label])
    for entry in extracted_data:
        final_ws.append(entry)

    # Requirement: Generate the bar or line chart
    chart = BarChart() if chart_type.lower() == 'bar' else LineChart()
    
    # Define references for values and labels
    # Values = Column B (2); Dates = Column A (1)
    values = Reference(final_ws, min_col=2, min_row=1, max_row=len(extracted_data) + 1)
    dates = Reference(final_ws, min_col=1, min_row=2, max_row=len(extracted_data) + 1)

    chart.add_data(values, titles_from_data=True)
    chart.set_categories(dates)

    # Requirement: Label axes and title (<student ID> <current date>)
    chart.x_axis.title = "Date"
    chart.y_axis.title = y_axis_label
    current_date = datetime.now().strftime("%m/%d/%Y")
    chart.title = f"Gerbla7878 {current_date}"

    final_ws.add_chart(chart, "D2")
    final_wb.save("final.xlsx")
    print("\nSUCCESS: 'final.xlsx' has been created.")

# Function: generateReport
# Description: Asks user for chart type and calls createChart.
# Argument: csv_path (str) - File path passed from main menu.
# Returns: None
def generateReport(csv_path):
    print("\n--- Generate Report ---")
    g_type = input("Enter graph type (line or bar): ").strip().lower()
    
    # Requirement: Call createChart and pass appropriate arguments
    createChart(csv_path, g_type)

# Updated Main Menu
def main_menu():
    # IMPORTANT: Use the 'r' prefix for Windows paths to avoid backslash errors
    source_file = r"C:\PythonFiles\final1.xlsx"

    while True:
        print("\n--- Main Menu ---")
        print("1. Process Data")
        print("2. Conversion")
        print("3. Generate Report")
        print("4. Exit")
        
        choice = input("Selection: ")
        if choice == "3":
            generateReport(source_file)
        elif choice == "4":
            print("Exiting.")
            break

# CRITICAL: This line makes the code actually run when you open the file
if __name__ == "__main__":
    main_menu()


